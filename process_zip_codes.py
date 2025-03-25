import pandas as pd
import re
import openpyxl
import os

def process_zip_codes(input_file="sample_data.xlsx", sheet_name=0, zip_column=None):
    """
    Creates a new sheet with only the ZIP column formatted as text.
    """
    try:
        #read the Excel file
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        #find the ZIP column if not specified
        if zip_column is None:
            home_zip_columns = [col for col in df.columns if 'home zip' in str(col).lower()]
            if home_zip_columns:
                zip_column = home_zip_columns[0]
            else:
                #taking the zip col but not the one with work
                zip_columns = [col for col in df.columns
                                if 'zip' in str(col).lower()
                                and 'work' not in str(col).lower()]
                if zip_columns:
                    zip_column = zip_columns[0]
                else:
                    raise ValueError("No suitable ZIP column found.")
        
        #create a dataframe with only the ZIP column
        zip_df = df[[zip_column]].copy()
        
        #process ZIP codes
        def format_zip(x):
            if pd.isna(x) or str(x).strip() == '':
                return ''
            
            #Extract digits only, this can be excluded as well but if excluded this will throw error or just give an invalid value in the quest report
            digits = re.sub(r'\D', '', str(x))
            
            #remove if less than 5 digits
            if len(digits) < 5:
                return ''
                
            #If more than 5 digits, trim to first 5
            trimmed_digits = digits[:5]
            
            #get rif of the obvious invalid values
            if trimmed_digits == "00000":
                return ''
                
            return trimmed_digits
        
        #apply the format zip function to every single value in the zip column
        zip_df[zip_column] = zip_df[zip_column].apply(format_zip)
        
        #remove rows with empty or invalid ZIP codes from the df object which is getting manipulated in memory
        zip_df = zip_df[zip_df[zip_column] != '']
        
        #till this point the original data is untouched
        #Create a new sheet name
        new_sheet_name = "ZIPs_Only"
        
        #save to a new sheet
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            zip_df.to_excel(writer, sheet_name=new_sheet_name, index=False)
        
        #Apply text formatting and create named range
        ##this helps in getting reference to the new sheet that was just created
        wb = openpyxl.load_workbook(input_file)
        #this is the sheet we are working with
        ws = wb[new_sheet_name]
        
        #Format as text for all cells including header
        for row in range(1, ws.max_row + 1):
            cell = ws[f"A{row}"]
            cell.number_format = '@'  #to convert into text format
        
        #Create a named range 'EE' for the ZIP column, DefinedName is used to create named ranges
        # Create a named range 'EE' for the ZIP column
        try:
            range_name = "EE"
            range_reference = f"'{new_sheet_name}'!$A$1:$A${ws.max_row}"
            
            # Remove existing name if present
            if range_name in wb.defined_names:
                del wb.defined_names[range_name]
            
            # Create the defined name using the dictionary-style assignment
            # This is more compatible across different openpyxl versions
            from openpyxl.workbook.defined_name import DefinedName
            defined_name = DefinedName(name=range_name, attr_text=range_reference)
            wb.defined_names[range_name] = defined_name
            print(f"Created named range 'EE' for column in {input_file}")
        except Exception as e:
            print(f"Error creating named range: {str(e)}")
            # Use VBA approach as final fallback
            try:
                # Create a VBA macro sheet with instructions
                vba_sheet_name = "VBA_Helper"
                if vba_sheet_name in wb.sheetnames:
                    wb.remove(wb[vba_sheet_name])
                    
                vba_sheet = wb.create_sheet(vba_sheet_name)
                vba_sheet["A1"] = "Sub CreateNamedRange()"
                vba_sheet["A2"] = f"    ThisWorkbook.Names.Add Name:=\"EE\", RefersTo:=\"='{new_sheet_name}'!$A$1:$A${ws.max_row}\""
                vba_sheet["A3"] = "End Sub"
                print(f"Created VBA Helper sheet. Run the macro to create the named range.")
            except:
                print("Could not create named range or helper. Please create it manually in Excel.")
                #Save the workbook
        wb.save(input_file)
        return True
    
    except Exception as e:
        return False
if __name__ == "__main__":
    #gets all the excel files in the current directory
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx') or f.endswith('.xls')]
    if excel_files:
        #processes the first excel file found in the list
        for file in excel_files:
            print(f"Processing {file}...")
            process_zip_codes(file)
    else:
        file_name = input("Enter Excel file name: ")
        if file_name:
            process_zip_codes(file_name)